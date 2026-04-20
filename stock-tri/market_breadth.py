"""
Market Breadth 트래커 (올인원) — Premium Edition
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
이 파일 하나만 실행하면:
  ① 엑셀 파일이 없으면 자동 생성 (차트 포함)
  ② 데이터를 가져와 자동 기록
  ③ KOSPI / KOSDAQ 분리 + 월별 시트 + 프리미엄 차트

설치:  pip install openpyxl finance-datareader pandas
실행:  python market_breadth.py
웹 UI:  streamlit run market_breadth_app.py  (첫 로드 시 캐시·데이터를 자동으로 불러와 차트 표시)
자동화: Windows 작업 스케줄러 or crontab → 매일 15:40 실행

© 올투스탁랩 ALLTOO STOCK LAB
"""

import pandas as pd
from datetime import datetime, timedelta
from pathlib import Path
import sys
import time

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.chart.marker import Marker
from openpyxl.chart.series import SeriesLabel
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.utils import get_column_letter

EXCEL_FILE = Path(__file__).parent / "market_breadth_tracker.xlsx"

C_NAVY = '0C1B2E'
C_GOLD = 'D4A843'
C_TEAL = '0D9488'
C_RED = 'DC2626'
C_ORANGE = 'F59E0B'
C_GREEN = '16A34A'
C_GRAY = '6B7280'
C_BLUE = '3B82F6'
C_WHITE = 'FFFFFF'
C_INPUT_BG = 'EFF6FF'


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  1. 데이터 수집 (캐시 지원)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━

CACHE_DIR = Path(__file__).parent / "breadth_cache"

def save_cache(name, df):
    CACHE_DIR.mkdir(exist_ok=True)
    df.to_pickle(CACHE_DIR / f"{name}.pkl")

def load_cache(name):
    path = CACHE_DIR / f"{name}.pkl"
    if path.exists():
        return pd.read_pickle(path)
    return None

def get_market_data(days=201):
    import FinanceDataReader as fdr

    today = datetime.today()
    end_str = today.strftime("%Y-%m-%d")

    # 캐시 확인
    kospi_cached = load_cache("kospi_pivot")
    kosdaq_cached = load_cache("kosdaq_pivot")
    kospi_idx_cached = load_cache("kospi_index")
    kosdaq_idx_cached = load_cache("kosdaq_index")

    has_cache = kospi_cached is not None and kosdaq_cached is not None

    if has_cache:
        last_date = kospi_cached.index[-1]
        if last_date >= end_str:
            print("  이미 최신 데이터입니다. 캐시 사용.\n", flush=True)
            return {
                'kospi_pivot': kospi_cached.iloc[-days:],
                'kosdaq_pivot': kosdaq_cached.iloc[-days:] if kosdaq_cached is not None else None,
                'kospi_index': kospi_idx_cached.iloc[-days:] if kospi_idx_cached is not None else None,
                'kosdaq_index': kosdaq_idx_cached.iloc[-days:] if kosdaq_idx_cached is not None else None,
            }
        # 증분 업데이트: 마지막 날짜 다음날부터
        next_day = (datetime.strptime(last_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
        print(f"  캐시 발견! 마지막 데이터: {last_date}", flush=True)
        print(f"  {next_day} ~ {end_str} 신규 데이터 확인 중...\n", flush=True)
        start_str = next_day

        # 먼저 지수로 신규 거래일이 있는지 확인 (빠른 체크)
        try:
            test_idx = fdr.DataReader('KS11', start_str, end_str)
            if test_idx.empty:
                print("  신규 거래일이 없습니다. (주말/공휴일)", flush=True)
                print("  기존 캐시로 엑셀을 생성합니다.\n", flush=True)
                return {
                    'kospi_pivot': kospi_cached.iloc[-days:],
                    'kosdaq_pivot': kosdaq_cached.iloc[-days:] if kosdaq_cached is not None else None,
                    'kospi_index': kospi_idx_cached.iloc[-days:] if kospi_idx_cached is not None else None,
                    'kosdaq_index': kosdaq_idx_cached.iloc[-days:] if kosdaq_idx_cached is not None else None,
                }
            new_dates = len(test_idx)
            print(f"  신규 거래일 {new_dates}일 발견! 데이터 수집합니다. (1~2분 소요)\n", flush=True)
        except Exception:
            print("  신규 데이터 확인 실패. 기존 캐시를 사용합니다.\n", flush=True)
            return {
                'kospi_pivot': kospi_cached.iloc[-days:],
                'kosdaq_pivot': kosdaq_cached.iloc[-days:] if kosdaq_cached is not None else None,
                'kospi_index': kospi_idx_cached.iloc[-days:] if kospi_idx_cached is not None else None,
                'kosdaq_index': kosdaq_idx_cached.iloc[-days:] if kosdaq_idx_cached is not None else None,
            }
    else:
        start = today - timedelta(days=days * 2)
        start_str = start.strftime("%Y-%m-%d")
        print(f"최초 실행 — 전체 데이터 수집 (약 15~20분 소요)\n", flush=True)

    # 지수 데이터
    print("  지수 데이터 수집 중...", flush=True)
    try:
        kospi_idx_new = fdr.DataReader('KS11', start_str, end_str)
        kosdaq_idx_new = fdr.DataReader('KQ11', start_str, end_str)
        print("  KOSPI/KOSDAQ 지수 완료", flush=True)
    except Exception as e:
        print(f"  [경고] 지수 조회 실패: {e}", flush=True)
        kospi_idx_new = pd.DataFrame()
        kosdaq_idx_new = pd.DataFrame()

    # 종목 목록
    print("  종목 목록 조회 중...", flush=True)
    try:
        kospi_list = fdr.StockListing('KOSPI')
        kosdaq_list = fdr.StockListing('KOSDAQ')
    except Exception as e:
        print(f"  [오류] 종목 목록 실패: {e}", flush=True)
        return None

    def extract_codes(df_list):
        """보통주만 추출 (ETF, ETN, 우선주, 스팩 제외)"""
        codes = []
        # 종목코드 컬럼 찾기
        code_col = None
        for cn in ['Code', 'Symbol', 'Ticker']:
            if cn in df_list.columns:
                code_col = cn
                break
        if code_col is None:
            code_col = df_list.columns[0]

        # 종목명 컬럼 찾기
        name_col = None
        for cn in ['Name', '종목명', '종목', 'SecurityName']:
            if cn in df_list.columns:
                name_col = cn
                break

        # ETF/ETN 브랜드 키워드
        etf_keywords = ['ETF', 'ETN', 'KODEX', 'TIGER', 'KBSTAR', 'ARIRANG',
                        'SOL', 'ACE', 'KOSEF', 'HANARO', 'FOCUS', 'TIMEFOLIO',
                        'PLUS', 'VITA', 'BNK', 'WOORI', 'RISE', 'TREX']

        for idx, row in df_list.iterrows():
            c = str(row[code_col]).strip()
            if len(c) != 6 or not c.isdigit():
                continue
            # 우선주 제외 (끝자리가 0이 아닌 것)
            if c[-1] != '0':
                continue
            # 종목명으로 ETF/ETN/스팩 제외
            if name_col is not None:
                name = str(row[name_col]).upper()
                if '스팩' in name or 'SPAC' in name:
                    continue
                if any(kw in name for kw in etf_keywords):
                    continue
            codes.append(c)
        return codes

    kospi_tickers = extract_codes(kospi_list)
    kosdaq_tickers = extract_codes(kosdaq_list)
    print(f"  KOSPI: {len(kospi_tickers)}개 / KOSDAQ: {len(kosdaq_tickers)}개 (보통주만)", flush=True)
    print(f"  (ETF, ETN, 우선주, 스팩 제외)\n", flush=True)

    def fetch_closes(tickers, label):
        closes = {}
        failed = 0
        t0 = time.time()
        for i, t in enumerate(tickers):
            try:
                df = fdr.DataReader(t, start_str, end_str)
                if not df.empty and 'Close' in df.columns:
                    closes[t] = df['Close']
            except Exception:
                failed += 1
            done = i + 1
            # 처음 20개 종목에서 성공이 0이면 조기 종료
            if done == 20 and len(closes) == 0:
                print(f"  [{label}] 신규 데이터 없음 — 조기 종료", flush=True)
                return closes, failed
            if done <= 3 or done % 50 == 0:
                elapsed = time.time() - t0
                rate = elapsed / done
                mins = int(rate * (len(tickers) - done) // 60)
                print(f"  [{label}] {done}/{len(tickers)} ({len(closes)}개 성공) - 약 {mins}분 남음", flush=True)
            time.sleep(0.05)
        return closes, failed

    print("  KOSPI 종목 수집...", flush=True)
    kospi_closes, f1 = fetch_closes(kospi_tickers, "KOSPI")
    print(f"  KOSPI 완료: {len(kospi_closes)}개 (실패: {f1})\n", flush=True)

    print("  KOSDAQ 종목 수집...", flush=True)
    kosdaq_closes, f2 = fetch_closes(kosdaq_tickers, "KOSDAQ")
    print(f"  KOSDAQ 완료: {len(kosdaq_closes)}개 (실패: {f2})\n", flush=True)

    def make_pivot(closes_dict):
        if not closes_dict:
            return None
        pivot = pd.DataFrame(closes_dict)
        pivot.index = pivot.index.strftime("%Y-%m-%d")
        pivot = pivot.sort_index()
        return pivot

    def clean_index(idx_df):
        if idx_df.empty:
            return None
        idx_df = idx_df[['Close']].copy()
        idx_df.index = idx_df.index.strftime("%Y-%m-%d")
        idx_df = idx_df.sort_index()
        return idx_df

    kospi_new = make_pivot(kospi_closes)
    kosdaq_new = make_pivot(kosdaq_closes)
    kospi_idx_clean = clean_index(kospi_idx_new)
    kosdaq_idx_clean = clean_index(kosdaq_idx_new)

    # 신규 데이터가 없으면 캐시 그대로 사용
    if kospi_new is None and kosdaq_new is None and has_cache:
        print("  신규 데이터 없음. 기존 캐시로 엑셀을 생성합니다.\n", flush=True)
        return {
            'kospi_pivot': kospi_cached.iloc[-days:],
            'kosdaq_pivot': kosdaq_cached.iloc[-days:] if kosdaq_cached is not None else None,
            'kospi_index': kospi_idx_cached.iloc[-days:] if kospi_idx_cached is not None else None,
            'kosdaq_index': kosdaq_idx_cached.iloc[-days:] if kosdaq_idx_cached is not None else None,
        }

    # 캐시와 병합
    def merge_data(cached, new):
        if cached is not None and new is not None:
            merged = pd.concat([cached, new])
            merged = merged[~merged.index.duplicated(keep='last')]
            return merged.sort_index()
        return new if new is not None else cached

    kospi_pivot = merge_data(kospi_cached, kospi_new)
    kosdaq_pivot = merge_data(kosdaq_cached, kosdaq_new)
    kospi_index = merge_data(kospi_idx_cached, kospi_idx_clean)
    kosdaq_index = merge_data(kosdaq_idx_cached, kosdaq_idx_clean)

    # 캐시 저장
    print("  캐시 저장 중...", flush=True)
    if kospi_pivot is not None:
        save_cache("kospi_pivot", kospi_pivot)
    if kosdaq_pivot is not None:
        save_cache("kosdaq_pivot", kosdaq_pivot)
    if kospi_index is not None:
        save_cache("kospi_index", kospi_index)
    if kosdaq_index is not None:
        save_cache("kosdaq_index", kosdaq_index)
    print("  캐시 저장 완료!\n", flush=True)

    return {
        'kospi_pivot': kospi_pivot.iloc[-days:] if kospi_pivot is not None else None,
        'kosdaq_pivot': kosdaq_pivot.iloc[-days:] if kosdaq_pivot is not None else None,
        'kospi_index': kospi_index.iloc[-days:] if kospi_index is not None else None,
        'kosdaq_index': kosdaq_index.iloc[-days:] if kosdaq_index is not None else None,
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  2. Breadth 계산
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def calc_breadth(pivot_df, label="", progress_callback=None):
    print(f"\n  [{label}] Breadth 계산 중...", flush=True)
    # 오류 데이터 제거 (공휴일, 비거래일 등)
    skip_dates = {"2025-06-17"}
    records = []
    for i in range(len(pivot_df)):
        date_str = pivot_df.index[i]
        if date_str in skip_dates:
            continue
        today_p = pivot_df.iloc[i].dropna()
        # 종목수가 너무 적으면 비거래일로 판단하여 스킵
        if len(today_p) < 100:
            continue
        row = {"date": date_str}

        for period in [200, 50, 20]:
            ma = pivot_df.iloc[max(0, i - period + 1):i + 1].mean()
            common = today_p.index.intersection(ma.dropna().index)
            above = int((today_p[common] >= ma[common]).sum())
            below = int((today_p[common] < ma[common]).sum())
            total = above + below
            row[f"ma{period}_above"] = above
            row[f"ma{period}_below"] = below
            row[f"ma{period}_pct"] = round(above / total * 100, 1) if total > 0 else 0

        lookback = min(250, i + 1)
        high_52 = pivot_df.iloc[max(0, i - lookback + 1):i + 1].max()
        low_52 = pivot_df.iloc[max(0, i - lookback + 1):i + 1].min()
        ch = today_p.index.intersection(high_52.dropna().index)
        cl = today_p.index.intersection(low_52.dropna().index)
        row["nh"] = int((today_p[ch] >= high_52[ch]).sum())
        row["nl"] = int((today_p[cl] <= low_52[cl]).sum())
        records.append(row)

        if (i + 1) % 20 == 0 or i == len(pivot_df) - 1:
            print(f"    {i+1}/{len(pivot_df)}일 완료", flush=True)
            if progress_callback is not None:
                progress_callback(i + 1, len(pivot_df), label)
    return records


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  3. 차트 헬퍼
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def style_chart_title(ch, color=C_NAVY):
    if ch.title:
        ch.title.txPr = RichText(
            p=[Paragraph(pPr=ParagraphProperties(
                defRPr=CharacterProperties(sz=1000, b=True, solidFill=color)
            ))]
        )


def make_index_chart(mkt, ws, start_r, max_r):
    ch = LineChart()
    ch.title = f"{mkt} Index"
    ch.style = 2
    ch.width, ch.height = 45, 10
    ch.y_axis.numFmt = '#,##0'
    ch.legend = None
    ch.x_axis.tickLblPos = 'low'
    ch.x_axis.delete = False
    cats = Reference(ws, min_col=14, min_row=start_r, max_row=max_r)
    vals = Reference(ws, min_col=2, min_row=start_r, max_row=max_r)
    ch.add_data(vals)
    ch.set_categories(cats)
    s = ch.series[0]
    s.tx = SeriesLabel(v="지수종가")
    clr = C_RED if mkt == "KOSPI" else C_TEAL
    s.graphicalProperties.line.solidFill = clr
    s.graphicalProperties.line.width = 18000
    s.smooth = False
    s.marker = Marker(symbol='circle', size=3)
    s.marker.graphicalProperties.solidFill = clr
    style_chart_title(ch)
    return ch


def make_ma_chart(mkt, ws, start_r, max_r):
    ch = LineChart()
    ch.title = f"MA Breadth (20ema 주황 / 50sma 초록 / 200sma 회색) - {mkt}"
    ch.style = 2
    ch.width, ch.height = 45, 12
    ch.y_axis.title = "비율 (%)"
    ch.y_axis.scaling.min = 0
    ch.y_axis.scaling.max = 100
    ch.y_axis.numFmt = '0"%"'
    ch.x_axis.tickLblPos = 'low'
    ch.x_axis.delete = False
    cats = Reference(ws, min_col=14, min_row=start_r, max_row=max_r)
    for col in [5, 8, 11]:
        vals = Reference(ws, min_col=col, min_row=start_r, max_row=max_r)
        ch.add_data(vals)
    ch.set_categories(cats)
    names = ["200sma", "50sma", "20ema"]
    colors = [C_GRAY, C_GREEN, C_ORANGE]
    widths = [18000, 20000, 22000]
    for i in range(3):
        ch.series[i].tx = SeriesLabel(v=names[i])
        ch.series[i].graphicalProperties.line.solidFill = colors[i]
        ch.series[i].graphicalProperties.line.width = widths[i]
        ch.series[i].smooth = False
    style_chart_title(ch)
    return ch


def make_nhnl_chart(mkt, ws, start_r, max_r):
    ch = LineChart()
    ch.title = f"{mkt} 52Weeks : NH-NL Breadth"
    ch.style = 2
    ch.width, ch.height = 45, 14
    ch.y_axis.title = "종목수"
    ch.x_axis.tickLblPos = 'low'
    ch.x_axis.delete = False
    cats = Reference(ws, min_col=14, min_row=start_r, max_row=max_r)
    for col in [12, 13]:
        vals = Reference(ws, min_col=col, min_row=start_r, max_row=max_r)
        ch.add_data(vals)
    ch.set_categories(cats)

    # 신고가: 빨간 실선 + 숫자 라벨
    s0 = ch.series[0]
    s0.tx = SeriesLabel(v="52주 신고가")
    s0.graphicalProperties.line.solidFill = C_RED
    s0.graphicalProperties.line.width = 20000
    s0.marker = Marker(symbol='circle', size=4)
    s0.marker.graphicalProperties.solidFill = C_RED
    s0.smooth = False
    s0.dLbls = DataLabelList()
    s0.dLbls.showVal = True
    s0.dLbls.showCatName = False
    s0.dLbls.showSerName = False
    s0.dLbls.numFmt = '#,##0'
    s0.dLbls.dLblPos = 't'
    s0.dLbls.txPr = RichText(
        p=[Paragraph(pPr=ParagraphProperties(
            defRPr=CharacterProperties(sz=650, b=True, solidFill=C_RED)
        ))]
    )

    # 신저가: 파란 점선 + 숫자 라벨
    s1 = ch.series[1]
    s1.tx = SeriesLabel(v="52주 신저가")
    s1.graphicalProperties.line.solidFill = C_BLUE
    s1.graphicalProperties.line.width = 18000
    s1.graphicalProperties.line.dashStyle = 'dash'
    s1.marker = Marker(symbol='circle', size=4)
    s1.marker.graphicalProperties.solidFill = C_BLUE
    s1.smooth = False
    s1.dLbls = DataLabelList()
    s1.dLbls.showVal = True
    s1.dLbls.showCatName = False
    s1.dLbls.showSerName = False
    s1.dLbls.numFmt = '#,##0'
    s1.dLbls.dLblPos = 'b'
    s1.dLbls.txPr = RichText(
        p=[Paragraph(pPr=ParagraphProperties(
            defRPr=CharacterProperties(sz=650, b=True, solidFill=C_BLUE)
        ))]
    )

    style_chart_title(ch)
    return ch


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  4. 엑셀 생성
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def build_excel(data, kospi_records, kosdaq_records, excel_path=None):
    out_file = excel_path if excel_path is not None else EXCEL_FILE
    wb = Workbook()

    hdr_fill = PatternFill('solid', fgColor=C_NAVY)
    hdr_font = Font(name='맑은 고딕', bold=True, color=C_WHITE, size=10)
    gold_fill = PatternFill('solid', fgColor=C_GOLD)
    gold_font = Font(name='맑은 고딕', bold=True, color=C_NAVY, size=10)
    data_font = Font(name='맑은 고딕', size=9)
    blue_font = Font(name='맑은 고딕', size=9, color='0000FF')
    thin = Border(bottom=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'))
    input_fill = PatternFill('solid', fgColor=C_INPUT_BG)

    headers = [
        "날짜", "지수종가",
        "200일 상회", "200일 하회", "200일 비율(%)",
        "50일 상회", "50일 하회", "50일 비율(%)",
        "20일 상회", "20일 하회", "20일 비율(%)",
        "52주 신고가", "52주 신저가", "차트날짜",
    ]
    widths = [12, 10, 11, 11, 12, 11, 11, 12, 11, 11, 12, 11, 11, 7]

    # ── KOSPI / KOSDAQ 데이터 시트 ──
    first_sheet = True
    for mkt, records, idx_data in [
        ("KOSPI", kospi_records, data.get('kospi_index')),
        ("KOSDAQ", kosdaq_records, data.get('kosdaq_index')),
    ]:
        if first_sheet:
            ws = wb.active
            ws.title = f"{mkt}_데이터"
            first_sheet = False
        else:
            ws = wb.create_sheet(f"{mkt}_데이터")

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = gold_fill if col in (5, 8, 11) else hdr_fill
            c.font = gold_font if col in (5, 8, 11) else hdr_font
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 28

        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        for ri, rec in enumerate(records, 2):
            date_val = datetime.strptime(rec["date"], "%Y-%m-%d").date()
            ws.cell(row=ri, column=1, value=date_val).number_format = 'YYYY-MM-DD'
            ws.cell(row=ri, column=1).font = data_font

            # 차트용 날짜 (M/D 형식 텍스트)
            ws.cell(row=ri, column=14, value=f"{date_val.month}/{date_val.day}")
            ws.cell(row=ri, column=14).font = Font(name='맑은 고딕', size=8, color=C_GRAY)
            ws.cell(row=ri, column=14).alignment = Alignment(horizontal='center')

            if idx_data is not None and rec["date"] in idx_data.index:
                ws.cell(row=ri, column=2, value=round(float(idx_data.loc[rec["date"], "Close"]), 2)).font = data_font
                ws.cell(row=ri, column=2).number_format = '#,##0.00'

            for col, key in [(3, "ma200_above"), (4, "ma200_below"),
                             (6, "ma50_above"), (7, "ma50_below"),
                             (9, "ma20_above"), (10, "ma20_below"),
                             (12, "nh"), (13, "nl")]:
                c = ws.cell(row=ri, column=col, value=rec[key])
                c.font = blue_font
                c.fill = input_fill
                c.alignment = Alignment(horizontal='center')

            for col, key in [(5, "ma200_pct"), (8, "ma50_pct"), (11, "ma20_pct")]:
                c = ws.cell(row=ri, column=col, value=rec[key])
                c.number_format = '0.0'
                c.font = data_font
                c.alignment = Alignment(horizontal='center')

            for col in range(1, 15):
                ws.cell(row=ri, column=col).border = thin

        ws.freeze_panes = 'A2'

    # ── 차트 시트 (KOSPI 왼쪽, KOSDAQ 오른쪽) ──
    wc = wb.create_sheet("차트_전체")
    wc.cell(row=1, column=1, value="Market Breadth : Stocks Above Key Moving Averages").font = Font(name='맑은 고딕', bold=True, size=18, color=C_NAVY)
    wc.cell(row=2, column=1, value=f"© {datetime.now().year} 올투스탁랩 ALLTOO STOCK LAB  |  최근 업데이트: {datetime.now().strftime('%Y-%m-%d')}").font = Font(name='맑은 고딕', size=9, color=C_GRAY)

    for ci, mkt in enumerate(["KOSPI", "KOSDAQ"]):
        col_anchor = "A" if ci == 0 else "AG"
        ws_data = wb[f"{mkt}_데이터"]
        max_r = ws_data.max_row

        # 최근 3개월(~63거래일)만 차트에 표시
        chart_days = 63
        start_r = max(2, max_r - chart_days + 1)

        wc.add_chart(make_index_chart(mkt, ws_data, start_r, max_r), f"{col_anchor}4")
        wc.add_chart(make_ma_chart(mkt, ws_data, start_r, max_r), f"{col_anchor}20")
        wc.add_chart(make_nhnl_chart(mkt, ws_data, start_r, max_r), f"{col_anchor}38")

    # 하단 저작권
    wc.cell(row=56, column=1, value=f"© {datetime.now().year} 올투스탁랩 ALLTOO STOCK LAB — 무단 복제 및 배포를 금합니다.").font = Font(name='맑은 고딕', size=8, color=C_GOLD)

    # ── 월별 시트 ──
    all_months = sorted(set(r["date"][:7] for r in kospi_records + kosdaq_records))

    for month in all_months:
        ws_m = wb.create_sheet(month)
        ws_m.cell(row=1, column=1, value=f"Market Breadth — {month}").font = Font(name='맑은 고딕', bold=True, size=14, color=C_NAVY)
        ws_m.cell(row=2, column=1, value=f"© {datetime.now().year} 올투스탁랩 ALLTOO STOCK LAB").font = Font(name='맑은 고딕', size=8, color=C_GOLD)

        row = 4
        for mkt, records in [("KOSPI", kospi_records), ("KOSDAQ", kosdaq_records)]:
            month_recs = [r for r in records if r["date"].startswith(month)]
            if not month_recs:
                continue

            ws_m.cell(row=row, column=1, value=f"■ {mkt}").font = Font(name='맑은 고딕', bold=True, size=12, color=C_NAVY)
            row += 1

            m_hdrs = ["날짜", "200일(%)", "50일(%)", "20일(%)", "신고가", "신저가"]
            for ci, h in enumerate(m_hdrs, 1):
                c = ws_m.cell(row=row, column=ci, value=h)
                c.fill = hdr_fill
                c.font = hdr_font
                c.alignment = Alignment(horizontal='center')
            row += 1

            for rec in month_recs:
                dv = datetime.strptime(rec["date"], "%Y-%m-%d").date()
                ws_m.cell(row=row, column=1, value=dv).number_format = 'MM-DD'
                ws_m.cell(row=row, column=1).font = data_font
                ws_m.cell(row=row, column=2, value=rec["ma200_pct"]).font = data_font
                ws_m.cell(row=row, column=2).number_format = '0.0'
                ws_m.cell(row=row, column=3, value=rec["ma50_pct"]).font = data_font
                ws_m.cell(row=row, column=3).number_format = '0.0'
                ws_m.cell(row=row, column=4, value=rec["ma20_pct"]).font = data_font
                ws_m.cell(row=row, column=4).number_format = '0.0'
                ws_m.cell(row=row, column=5, value=rec["nh"]).font = blue_font
                ws_m.cell(row=row, column=6, value=rec["nl"]).font = Font(name='맑은 고딕', size=9, color=C_RED)
                for ci in range(1, 7):
                    ws_m.cell(row=row, column=ci).border = thin
                    ws_m.cell(row=row, column=ci).alignment = Alignment(horizontal='center')
                row += 1
            row += 2

        for ci, w in enumerate([12, 10, 10, 10, 10, 10], 1):
            ws_m.column_dimensions[get_column_letter(ci)].width = w

    # ── 사용법 시트 ──
    wi = wb.create_sheet("사용법")
    wi.column_dimensions['A'].width = 90
    t_font = Font(name='맑은 고딕', bold=True, size=14, color=C_NAVY)
    s_font = Font(name='맑은 고딕', bold=True, size=11, color=C_NAVY)
    b_font = Font(name='맑은 고딕', size=10)
    r_font = Font(name='맑은 고딕', size=10, color=C_RED)

    usage = [
        ("【 Market Breadth 트래커 — Premium Edition 】", t_font),
        (f"© {datetime.now().year} 올투스탁랩 ALLTOO STOCK LAB", Font(name='맑은 고딕', size=10, color=C_GOLD)),
        ("", b_font),
        ("■ 시트 구성", s_font),
        ("  KOSPI_데이터 / KOSDAQ_데이터: 원본 + 지수 + 비율", b_font),
        ("  차트_전체: 6개 차트 (지수 + MA Breadth + NH-NL × KOSPI/KOSDAQ)", b_font),
        ("  월별 시트 (YYYY-MM): 월간 데이터 요약", b_font),
        ("", b_font),
        ("■ 차트 색상", s_font),
        ("  200일 이평선: 회색 (장기)", b_font),
        ("  50일 이평선: 초록 (중기)", b_font),
        ("  20일 이평선: 주황 (단기)", b_font),
        ("  52주 신고가: 파란 점선  /  신저가: 빨간 실선", b_font),
        ("", b_font),
        ("【 해석 가이드 】", t_font),
        ("  비율 50%↑ = 강세  /  50%↓ = 약세", b_font),
        ("  20일 먼저 반등 → 50일 → 200일 순 회복 = 건강한 랠리", b_font),
        ("  지수 신고가 + 신고가 종목 증가 = Breadth Confirmation", b_font),
        ("  지수 신고가인데 신고가 종목 감소 = Breadth Divergence (위험!)", r_font),
    ]
    for i, (text, font) in enumerate(usage, 1):
        wi.cell(row=i, column=1, value=text).font = font

    wb.save(out_file)
    print(f"\n✅ 엑셀 저장: {out_file}", flush=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  실행
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def main():
    print("=" * 55, flush=True)
    print("  Market Breadth — Premium Edition", flush=True)
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", flush=True)
    print(f"  © 올투스탁랩 ALLTOO STOCK LAB", flush=True)
    print("=" * 55, flush=True)

    data = get_market_data(days=201)
    if data is None:
        sys.exit(1)

    kospi_recs = calc_breadth(data['kospi_pivot'], "KOSPI") if data['kospi_pivot'] is not None else []
    kosdaq_recs = calc_breadth(data['kosdaq_pivot'], "KOSDAQ") if data['kosdaq_pivot'] is not None else []

    if not kospi_recs and not kosdaq_recs:
        print("\n[오류] 계산된 데이터가 없습니다.", flush=True)
        sys.exit(1)

    for label, recs in [("KOSPI", kospi_recs), ("KOSDAQ", kosdaq_recs)]:
        if recs:
            L = recs[-1]
            print(f"\n📊 [{label}] {L['date']}:", flush=True)
            print(f"  200일: {L['ma200_pct']}% | 50일: {L['ma50_pct']}% | 20일: {L['ma20_pct']}%", flush=True)
            print(f"  52주 신고가: {L['nh']} / 신저가: {L['nl']}", flush=True)

    build_excel(data, kospi_recs, kosdaq_recs)
    print("\n✅ 완료!", flush=True)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        import traceback
        print(f"\n[오류 발생] {e}", flush=True)
        traceback.print_exc()
    input("\n아무 키나 누르면 종료됩니다...")
